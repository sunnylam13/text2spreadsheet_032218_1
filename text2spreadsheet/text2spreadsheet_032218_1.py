# -*- coding: utf-8 -*-

#! python3

# USAGE
# python3 text2spreadsheet_032218_1.py FOLDERPATH
# python3 text2spreadsheet_032218_1.py "../tests/testFolder1"

import openpyxl, sys, os, re

try:
	from openpyxl.cell import column_index_from_string,get_column_letter
except ImportError:
	from openpyxl.utils import column_index_from_string,get_column_letter

import logging
logging.basicConfig(level=logging.DEBUG, format=" %(asctime)s - %(levelname)s - %(message)s")
# logging.disable(logging.CRITICAL)

#####################################
# PARSE COMMAND LINE
#####################################

# get the folder name / folder path containing the text files that will be put into the spreadsheet

folder_to_process = sys.argv[1] # this should be the filename/file path

#####################################
# END PARSE COMMAND LINE
#####################################

#####################################
# REGEX
#####################################

# create a regex statement to match `user_file_ext_input`
# https://regexr.com/3kvi4
# re.compile should turn a raw string into current regex language so you can skip creating the formula sort of...

user_file_ext_input = '.txt' # for this program we only want `.txt` or text files so we'll set this now instead of taking input

file_type_regex1 = re.compile(user_file_ext_input + "$")
# print(file_type_regex1) # for testing
# print(file_type_regex1.search("testTextA1.txt")) # for testing

#####################################
# END REGEX
#####################################

#####################################
# SCAN FOLDERS/FILES
#####################################

# analyze the folder
# build a list of all the text files in the folder
# construct the list so that it is file paths to each text file

# get the absolute file path of the current working directory of program
abs_cwd_file_path = os.path.abspath('.') # set the destination file path to be the current working directory or cwd

# a list of all folders and subfolders to be analyzed
folder_path_list = [] # a list to hold all finalized folder paths (not folder names)

# a list of all files to be analyzed
file_path_list = [] # a list to hold all finalized folder paths (not folder names)

def scanFolder(foldername_path):
	# this function scans the parent folder and subfolders
	# it then adds them to a list so that its files can be scanned individually

	# `foldername_path` should actually be a string path to folder
	# `dirPath` - the directory path leading up to the folder's name (yet not including it), should be an absolute path I'd say
	
	# as we get deeper and deeper into subfolders it should add onto the folder's path string that we pass to it so accuracy should be maintained

	dirs = os.listdir(foldername_path) # list all files of any kind (i.e. all file and folder names)

	for file in dirs:
		# new_path = os.path.join(absPath,file) # creates a path to the file/folder
		new_path = os.path.join(foldername_path,file) # creates a path to the file/folder

		if os.path.isdir(new_path): #if the file is a folder
			folder_path_list.append(new_path) # add it to the list of folders with its full path name
		else:
			continue # otherwise skip and keep going

def scanFile(foldername_path,regex):
	# the file scanner that gets all of the files and pushes them into a list after we get the full string path to it
	
	dirs = os.listdir(foldername_path) # list all files of any kind (i.e. all file and folder names)

	for file in dirs:
		# new_path = os.path.join(absPath,file) # creates a path to the file/folder
		new_path = os.path.join(foldername_path,file) # creates a path to the file/folder

		if os.path.isfile(new_path) and regex.search(file): #if the file is a folder AND has regex match
			file_path_list.append(new_path) # add it to the list of folders with its full path name
		else:
			continue # otherwise skip and keep going

def generate_targets(user_folder_input,regexMatch):
	# run an initial scan of the upper level main folder tree
	# find subfolders if any
	scanFolder(user_folder_input)
	# find matching files or just files
	scanFile(user_folder_input,regexMatch)

	# then scan all the sub folders by cycling through folder_path_list until no more subfolders are added
	# this should keep going until no more subfolders are analyzed
	# then scan all the files by cycling through folder_path_list until no more subfolders are added/left
	for subfolder in folder_path_list:
		scanFolder(subfolder)
		scanFile(subfolder,regexMatch)

	logging.debug('folder_path_list is now:  ')
	logging.debug(folder_path_list)
	logging.debug('file_path_list is now:  ')
	logging.debug(file_path_list)

generate_targets(folder_to_process,file_type_regex1) # create paths to all files we want to process

#####################################
# END SCAN FOLDERS/FILES
#####################################

#####################################
# CREATE SPREADSHEET
#####################################

# analyze the files now stored in `file_path_list`
# the function should go through a list of text files (i.e. text file paths)
# then write them to a spreadsheet

def write_text_2_sheet(workbook,column_counter,line_list):
	# switch to active sheet / specific sheet
	sheet = workbook.active

	# convert the column_counter number into a column letter
	column_letter = get_column_letter(column_counter)
	logging.debug('The column letter to write is: %s' % (column_letter) )

	# now we iterate by row_number (within the same column)
	# the number of rows will be equal to the number of lines of text (i.e. len(line_list) or the length of said list) + 1 (to get the last number)
	# unfortunately need number range or would have used `for...in` loop
	for row_number in range(1,len(line_list)+1):
		logging.debug('The row number to write is: %s' % (str(row_number)) )
		# create the column and row coordinate
		cell_coordinate = column_letter + str(row_number)
		logging.debug('The cell_coordinate is: %s' % (cell_coordinate) )
		# set the value of the cell_coordinate
		logging.debug('The line_list index is: %s' % (row_number - 1) )
		logging.debug('The line_list value is: %s' % (line_list[row_number - 1]) )
		sheet[cell_coordinate] = line_list[row_number - 1] # we subtract 1 because lists start at n = 0 unlike spreadsheet labels which start at n = 1

def text2spreadsheet_generator(file_path_list):
	
	# create a new spreadsheet workbook to store the text
	nwb = openpyxl.Workbook()

	# set the counter used to assign column positioning
	column_counter = 1

	# for each text file in the `file_path_list`
	for text_file_path in file_path_list:

		# open the text file in `read` mode
		text_target = open(text_file_path,'r+')
		# go through it line by line, this generates a list of strings
		line_list = text_target.readlines()
		logging.debug('Text file readlines() results in:  ')
		logging.debug(line_list)
		
		# write the line(s) to the spreadsheet
		write_text_2_sheet(nwb,column_counter,line_list)
		logging.debug('Lines of text file have been written.')

		# close the text file you opened in `read` mode
		text_target.close()
		logging.debug('Text file has been closed.')

		# increment counter used to assign column positioning
		column_counter += 1
		logging.debug('column_counter incremented.')
	
	# save the new spreadsheet
	nwb.save('text2spreadsheetfinal.xlsx')
	logging.debug('Spreadsheet file saved.')


#####################################
# END CREATE SPREADSHEET
#####################################


#####################################
# EXECUTION
#####################################

text2spreadsheet_generator(file_path_list)

#####################################
# END EXECUTION
#####################################



